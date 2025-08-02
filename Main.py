import os
import csv
import uuid
import json
import subprocess
import sys
from datetime import datetime
import tkinter as tk
from tkinter import (
    Toplevel, Label, Entry, Button, StringVar, IntVar,
    Frame, PhotoImage, filedialog, Menu, messagebox, END,
    Scrollbar, Checkbutton
)
from tkinter import ttk
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from barcode import Code128
from fpdf import FPDF
from pathlib import Path
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont, ImageTk
import openpyxl
import shutil

import arabic_reshaper
from bidi.algorithm import get_display
import customtkinter as ctk
import pygame
import urllib.request
import shutil
import requests





# === CONFIG LICENCE ===
URL_VERIFICATION = "https://script.google.com/macros/s/AKfycbzwu7neoNhPi-tehESqABnyo6pwh43Vu8g0OfI_Y_s8fJTLWRduDlj2I4BPIPgSnSAvPg/exec"
FICHIER_LICENCE = "licence.json"

def get_machine_id():
    return hex(uuid.getnode())

def verifier_licence_locale():
    if not os.path.exists(FICHIER_LICENCE):
        return False
    try:
        with open(FICHIER_LICENCE, 'r') as f:
            data = json.load(f)
            return data.get("valide", False) and data.get("machine") == get_machine_id()
    except:
        return False

def enregistrer_licence_locale(cle):
    with open(FICHIER_LICENCE, 'w') as f:
        json.dump({"valide": True, "cle": cle, "machine": get_machine_id()}, f)

def verifier_licence_en_ligne(cle):
    try:
        response = requests.post(URL_VERIFICATION, json={
            "cle": cle,
            "machine": get_machine_id()
        }, timeout=10)
        data = response.json()
        if data.get("valide"):
            enregistrer_licence_locale(cle)
        return data
    except Exception as e:
        return {"valide": False, "message": "Erreur de connexion"}

def demander_activation(root):
    def valider():
        cle = champ_cle.get().strip()
        result = verifier_licence_en_ligne(cle)
        if result["valide"]:
            messagebox.showinfo("Activation réussie", result["message"])
            fenetre.destroy()
            root.destroy()
            os.system("python Main.py")
        else:
            messagebox.showerror("Erreur", result["message"])

    fenetre = Toplevel(root)
    fenetre.title("Activation")
    fenetre.geometry("350x120")
    Label(fenetre, text="Entrez votre clé de licence :").pack(pady=5)
    champ_cle = Entry(fenetre)
    champ_cle.pack(padx=10)
    Button(fenetre, text="Valider", command=valider).pack(pady=10)




VERSION_ACTUELLE = "1.0"

pygame.mixer.init()

def preparer_texte_arabe(texte):
    try:
        reshaped_text = arabic_reshaper.reshape(texte)
        return get_display(reshaped_text)
    except Exception as e:
        print(f"Erreur lors de la préparation du texte arabe : {e}")
        return texte


app = None

FICHIER_ELEVES = "liste_eleves.csv"
FICHIER_PRESENCES = "presences_cantine.csv"
FICHIER_CONFIG = "config.json"

def charger_config():
    if os.path.exists(FICHIER_CONFIG):
        try:
            with open(FICHIER_CONFIG, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            messagebox.showerror("Erreur de configuration", "Le fichier de configuration est corrompu. Utilisation des paramètres par défaut.")
            return {"mot_de_passe": "admin123"}
    return {"mot_de_passe": "admin123"}

def sauvegarder_config(config):
    try:
        with open(FICHIER_CONFIG, 'w', encoding='utf-8') as f:
            json.dump(config, indent=4)
    except IOError as e:
        messagebox.showerror("Erreur de sauvegarde", f"Impossible de sauvegarder le fichier de configuration : {e}")

if not os.path.exists(FICHIER_ELEVES):
    with open(FICHIER_ELEVES, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Nom", "Prénom", "Classe", "Formule"])

if not os.path.exists(FICHIER_PRESENCES):
    with open(FICHIER_PRESENCES, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Nom", "Prénom", "Classe", "Date", "Heure"])

def changer_mot_de_passe():
    def valider_changement():
        config = charger_config()

        ancien = ancien_mdp.get().strip()
        nouveau = nouveau_mdp.get().strip()

        if ancien != config["mot_de_passe"]:
            messagebox.showerror("Erreur", "Ancien mot de passe incorrect")
            return

        config["mot_de_passe"] = nouveau
        sauvegarder_config(config)

        global config_active
        config_active = config

        messagebox.showinfo("Succès", "Mot de passe changé. Vous pouvez maintenant vous connecter avec le nouveau mot de passe.")
        fen_changement.destroy()

    # Pass fenetre_mdp_instance as the parent for transient property
    # This assumes fenetre_mdp_instance is active when changer_mot_de_passe is called.
    # If called from elsewhere, this might need adjustment.
    parent_window = fenetre_mdp_instance if fenetre_mdp_instance and fenetre_mdp_instance.winfo_exists() else racine
    fen_changement = ctk.CTkToplevel(parent_window)
    fen_changement.title("Changer le mot de passe")
    fen_changement.geometry("350x200")

    # Make it transient to the parent window
    if parent_window:
        fen_changement.transient(parent_window)

    fen_changement.update_idletasks()
    largeur_fenetre = fen_changement.winfo_width()
    hauteur_fenetre = fen_changement.winfo_height()
    largeur_ecran = fen_changement.winfo_screenwidth()
    hauteur_ecran = fen_changement.winfo_screenheight()
    x = (largeur_ecran // 2) - (largeur_fenetre // 2)
    y = (hauteur_ecran // 2) - (hauteur_fenetre // 2)
    fen_changement.geometry(f"+{x}+{y}")

    # Make the window grab focus, ensuring it's on top
    fen_changement.grab_set()

    ctk.CTkLabel(fen_changement, text="Ancien mot de passe:    ", font=("Arial", 12)).pack(pady=(10, 0))
    ancien_mdp = ctk.CTkEntry(fen_changement, show="*", font=("Arial", 12))
    ancien_mdp.pack(pady=0)

    ctk.CTkLabel(fen_changement, text="Nouveau mot de passe:", font=("Arial", 12)).pack(pady=0)
    nouveau_mdp = ctk.CTkEntry(fen_changement, show="*", font=("Arial", 12))
    nouveau_mdp.pack(pady=0)

    ctk.CTkButton(fen_changement, text="Valider", font=("Arial", 12), command=valider_changement).pack(pady=(20, 0))

    # Wait for this window to close
    parent_window.wait_window(fen_changement)

fenetre_mdp_instance = None # Global variable to keep track of the admin password window instance
admin_button_reference = None # Global variable to hold the reference to the "Administrateur" button

def verifier_mot_de_passe():
    global fenetre_mdp_instance, admin_button_reference
    config_active = charger_config()

    # Si une instance existe déjà, on la met au premier plan
    if fenetre_mdp_instance and fenetre_mdp_instance.winfo_exists():
        fenetre_mdp_instance.lift()
        fenetre_mdp_instance.focus_force()
        return

    # Désactiver le bouton "Administrateur"
    if admin_button_reference:
        admin_button_reference.configure(state="disabled")

    # Variable pour suivre si une vérification est en cours
    verification_en_cours = False

    def verifier():
        nonlocal verification_en_cours
        # Empêcher les déclenchements multiples
        if verification_en_cours:
            return
            
        verification_en_cours = True
        
        config = charger_config()
        if mot_de_passe.get() == config["mot_de_passe"]:
            fenetre_mdp_instance.destroy()
            racine.destroy()
            lancer_application()
        else:
            messagebox.showerror("Erreur", "Mot de passe incorrect")
            # Réactiver après l'affichage du message
            fenetre_mdp_instance.after(100, lambda: setattr(verifier, 'verification_en_cours', False))
        
        verification_en_cours = False

    fenetre_mdp_instance = ctk.CTkToplevel()
    fenetre_mdp_instance.title("Accès Administrateur")
    fenetre_mdp_instance.transient(racine)
    fenetre_mdp_instance.geometry("350x180")

    # Centrer la fenêtre
    fenetre_mdp_instance.update_idletasks()
    largeur_fenetre = fenetre_mdp_instance.winfo_width()
    hauteur_fenetre = fenetre_mdp_instance.winfo_height()
    largeur_ecran = fenetre_mdp_instance.winfo_screenwidth()
    hauteur_ecran = fenetre_mdp_instance.winfo_screenheight()
    x = (largeur_ecran // 2) - (largeur_fenetre // 2)
    y = (hauteur_ecran // 2) - (hauteur_fenetre // 2)
    fenetre_mdp_instance.geometry(f"+{x}+{y}")

    # Widgets
    ctk.CTkLabel(fenetre_mdp_instance, text="Entrez le mot de passe", font=("Arial", 13)).pack(pady=(15, 0))
    
    mot_de_passe = ctk.CTkEntry(fenetre_mdp_instance, show="*", font=("Arial", 12), placeholder_text="Mot de passe")
    mot_de_passe.pack(pady=0)
    mot_de_passe.focus_set()  # Donner le focus au champ de mot de passe

    # Bouton Entrer
    btn_entrer = ctk.CTkButton(fenetre_mdp_instance, text="Entrer", font=("Arial", 12), command=verifier)
    btn_entrer.pack(pady=15)

    # Lier la touche Entrer UNIQUEMENT au champ de mot de passe
    mot_de_passe.bind('<Return>', lambda event: verifier())

    # Lien pour changer le mot de passe
    changer_mdp_lien = ctk.CTkLabel(fenetre_mdp_instance, text="Changer le mot de passe", 
                                  text_color="blue", cursor="hand2", font=("Arial", 12))
    changer_mdp_lien.pack(pady=0)
    changer_mdp_lien.bind("<Button-1>", lambda e: changer_mot_de_passe())

    # Rendre la fenêtre modale
    fenetre_mdp_instance.grab_set()
    
    # Nettoyage quand la fenêtre se ferme
    def on_close():
        fenetre_mdp_instance.unbind('<Return>')
        if admin_button_reference:
            admin_button_reference.configure(state="normal")
        fenetre_mdp_instance.destroy()
    
    fenetre_mdp_instance.protocol("WM_DELETE_WINDOW", on_close)
    
    racine.wait_window(fenetre_mdp_instance)
    
    
    
    
def verifier_mise_a_jour():
    try:
        # 🔗 Liens vers tes fichiers sur Google Drive (convertis en lien direct)
        lien_version = "https://drive.usercontent.google.com/download?id=1pmrGS3PZUUpQdHZ44tRUK772wwxo7DWE&export=download"
        lien_script = "https://drive.usercontent.google.com/download?id=1K013O2eTzqV9Wo7RV3EMH7boygiNHaZ2&export=download"

        # 1. Télécharger version.txt
        version_distance = None
        with urllib.request.urlopen(lien_version) as response:
            version_distance = response.read().decode('utf-8').strip()

        if not version_distance:
            messagebox.showerror("Erreur", "Impossible de lire la version en ligne.")
            return

        if version_distance == VERSION_ACTUELLE:
            messagebox.showinfo("Mise à jour", f"Mise à jour actuelle : {VERSION_ACTUELLE}\nAucune nouvelle mise à jour disponible.")
            return

        # 2. Nouvelle version disponible
        reponse = messagebox.askyesno(
            "Nouvelle mise à jour disponible",
            f"Mise à jour actuelle : {VERSION_ACTUELLE}\nNouvelle mise à jour disponible : {version_distance}\n\nSouhaitez-vous la télécharger ?"
        )

        if reponse:
            chemin_local = os.path.abspath(__file__)
            fichier_temporaire = "Main_temp.py"

            # Télécharger la nouvelle version dans un fichier temporaire
            urllib.request.urlretrieve(lien_script, fichier_temporaire)

            # Sauvegarder l'ancienne version
            sauvegarde = "Main_backup.py"
            shutil.copy2(chemin_local, sauvegarde)

            # Remplacer le fichier actuel
            shutil.move(fichier_temporaire, chemin_local)

            messagebox.showinfo("Mise à jour réussie", f"L'application a été mise à jour vers la version {version_distance}.\nRedémarre l'application pour voir les changements.")
        else:
            messagebox.showinfo("Annulé", "Mise à jour annulée par l'utilisateur.")

    except Exception as e:
        messagebox.showerror("Erreur", f"Échec de la mise à jour :\n{e}")

    
    
    
    

def ouvrir_fenetre_fiche():
    fen = ctk.CTkToplevel()
    fen.transient(app)        # La rattache à la fenêtre Menu Administrateur
    fen.grab_set()
    fen.title("Créer une fiche élève")
    fen.geometry("600x550")
    fen.resizable(False, False)

    font_label = ("Arial", 14)
    font_entry = ("Arial", 13)
    font_button = ("Arial", 13)
    font_check = ("Arial", 13)

    fen.update_idletasks()
    largeur = fen.winfo_width()
    hauteur = fen.winfo_height()
    x = (fen.winfo_screenwidth() // 2) - (largeur // 2)
    y = (fen.winfo_screenheight() // 2) - (hauteur // 2)
    fen.geometry(f"+{x}+{y}")

    main_frame = ctk.CTkFrame(fen)
    main_frame.pack(padx=20, pady=20, fill="both", expand=True)

    ctk.CTkLabel(main_frame, text="Nom", font=font_label).pack(pady=5)
    entry_nom = ctk.CTkEntry(main_frame, font=font_entry, width=300)
    entry_nom.pack(pady=5)

    ctk.CTkLabel(main_frame, text="Prénom", font=font_label).pack(pady=5)
    entry_prenom = ctk.CTkEntry(main_frame, font=font_entry, width=300)
    entry_prenom.pack(pady=5)

    ctk.CTkLabel(main_frame, text="Classe", font=font_label).pack(pady=5)
    entry_classe = ctk.CTkEntry(main_frame, font=font_entry, width=300)
    entry_classe.pack(pady=5)

    ctk.CTkLabel(main_frame, text="Formule", font=font_label).pack(pady=5)

    demi_pension_var = ctk.IntVar()
    externe_var = ctk.IntVar()

    frame_formule = ctk.CTkFrame(main_frame, fg_color="transparent")
    frame_formule.pack(pady=5)

    ctk.CTkCheckBox(frame_formule, text="Demi-pension", font=font_check,
                    variable=demi_pension_var).pack(side="left", padx=10)
    ctk.CTkCheckBox(frame_formule, text="Externe", font=font_check,
                    variable=externe_var).pack(side="left", padx=10)

    photo_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
    photo_frame.pack(pady=10)

    label_photo = ctk.CTkLabel(photo_frame, text="Aucune photo sélectionnée", font=font_label)
    label_photo.pack(pady=5)

    def choisir_photo():
        nonlocal photo_path
        photo_path = filedialog.askopenfilename(
            title="Choisir une photo",
            filetypes=[("Images", "*.png;*.jpg;*.jpeg")])
        if photo_path:
            label_photo.configure(text=os.path.basename(photo_path))
        else:
            label_photo.configure(text="Aucune photo sélectionnée")

    ctk.CTkButton(photo_frame, text="Choisir une photo", font=("Arial", 15),
              command=choisir_photo, width=180, height=35).pack(pady=5)


    btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
    btn_frame.pack(pady=20)

    ctk.CTkButton(btn_frame, text="Créer la fiche élève", font=("Arial", 15),
              width=250, height=45,
                  command=lambda: creer_fiche(fen, entry_nom, entry_prenom, entry_classe,
                                              demi_pension_var, externe_var, photo_path)).pack()

    photo_path = ""
    
    
def creer_fiche(fen, entry_nom, entry_prenom, entry_classe, demi_pension_var, externe_var, photo_path):
    nom = entry_nom.get().strip()
    prenom = entry_prenom.get().strip()
    classe = entry_classe.get().strip()

    if not nom or not prenom or not classe:
        messagebox.showerror("Erreur", "Tous les champs doivent être remplis.")
        return

    if demi_pension_var.get() and externe_var.get():
        messagebox.showerror("Erreur", "Veuillez sélectionner une seule formule (Demi-pension OU Externe).")
        return
    elif not demi_pension_var.get() and not externe_var.get():
        messagebox.showerror("Erreur", "Veuillez sélectionner une formule (Demi-pension ou Externe).")
        return

    if demi_pension_var.get():
        formule = "Demi-pension"
    else:
        formule = "Externe"

    identifiant = str(uuid.uuid4())[:8]

    dossier_fiches = "fiches_eleves"
    os.makedirs(dossier_fiches, exist_ok=True)

    dossier_eleve = os.path.join(dossier_fiches, f"{nom}_{prenom}")
    os.makedirs(dossier_eleve, exist_ok=True)

    chemin_codebarre = os.path.join(dossier_eleve, f"{identifiant}.png")
    Code128(identifiant, writer=ImageWriter()).save(chemin_codebarre[:-4])

    chemin_photo = os.path.join(dossier_eleve, f"{nom.lower()}_{prenom.lower()}.png")
    if photo_path and os.path.exists(photo_path):
        try:
            Image.open(photo_path).save(chemin_photo)
        except Exception as e:
            messagebox.showwarning("Erreur photo", f"Impossible de sauvegarder la photo : {e}")
            chemin_photo = None
    else:
        chemin_photo = None

    nom_pdf = os.path.join(dossier_eleve, f"fiche_{nom}_{prenom}.pdf")
    c = canvas.Canvas(nom_pdf, pagesize=A4)
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        font_path = os.path.join(os.path.dirname(__file__), "arial.ttf")
        if not os.path.exists(font_path):
            messagebox.showwarning("Police manquante", "Le fichier 'arial.ttf' est introuvable. Le texte arabe pourrait ne pas s'afficher correctement dans le PDF de la fiche.")
            c.setFont("Helvetica-Bold", 20)
            c.drawString(100, 780, "Fiche Élève")
            c.setFont("Helvetica", 14)
        else:
            pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
            c.setFont("ArabicFont", 20)
            c.drawString(100, 780, preparer_texte_arabe("Fiche Élève"))
            c.setFont("ArabicFont", 14)
            c.drawString(100, 740, preparer_texte_arabe(f"Nom : {nom}"))
            c.drawString(100, 720, preparer_texte_arabe(f"Prénom : {prenom}"))
            c.drawString(100, 700, preparer_texte_arabe(f"Classe : {classe}"))
            c.drawString(100, 680, f"Identifiant : {identifiant}")
            c.drawString(100, 660, preparer_texte_arabe(f"Formule : {formule}"))
    except Exception as e:
        print(f"Erreur lors de l'enregistrement de la police pour ReportLab: {e}")
        c.setFont("Helvetica-Bold", 20)
        c.drawString(100, 780, "Fiche Élève")
        c.setFont("Helvetica", 14)
        c.drawString(100, 740, f"Nom : {nom}")
        c.drawString(100, 720, f"Prénom : {prenom}")
        c.drawString(100, 700, f"Classe : {classe}")
        c.drawString(100, 680, f"Identifiant : {identifiant}")
        c.drawString(100, 660, f"Formule : {formule}")

    if chemin_photo and os.path.exists(chemin_photo):
        c.drawImage(chemin_photo, 400, 650, width=120, height=120)
    else:
        c.drawString(400, 700, preparer_texte_arabe("❌ Photo non trouvée"))
    c.drawImage(chemin_codebarre, 100, 500, width=450, height=150)
    c.save()

    badge = Image.new("RGB", (1181, 709), "white")
    draw = ImageDraw.Draw(badge)

    font_titre = None
    font_info = None
    try:
        font_path = os.path.join(os.path.dirname(__file__), "arial.ttf")
        if os.path.exists(font_path):
            font_titre = ImageFont.truetype(font_path, 60)
            font_info = ImageFont.truetype(font_path, 40)
        else:
            font_titre = ImageFont.load_default()
            font_info = ImageFont.load_default()
    except Exception as e:
        print(f"Erreur lors du chargement de la police pour PIL: {e}")
        font_titre = ImageFont.load_default()
        font_info = ImageFont.load_default()

    draw.text((480, 20), preparer_texte_arabe("Carte d'accès"), fill="black", font=font_titre)
    draw.text((40, 150), preparer_texte_arabe(f"Nom : {nom.upper()}"), fill="black", font=font_info)
    draw.text((40, 220), preparer_texte_arabe(f"Prénom : {prenom.upper()}"), fill="black", font=font_info)
    draw.text((40, 290), preparer_texte_arabe(f"Classe : {classe}"), fill="black", font=font_info)
    draw.text((40, 360), f"ID : {identifiant}", fill="black", font=font_info)
    draw.text((40, 430), preparer_texte_arabe(f"Formule : {formule}"), fill="black", font=font_info)

    if chemin_photo and os.path.exists(chemin_photo):
        photo = Image.open(chemin_photo).resize((320, 320))
        badge.paste(photo, (801, 150))
    else:
        draw.text((850, 150), preparer_texte_arabe(""), fill="black", font=font_info)

    if os.path.exists(chemin_codebarre):
        code_img = Image.open(chemin_codebarre).resize((750, 225))
        badge.paste(code_img, (215, 490))
    else:
        draw.text((540, 600), preparer_texte_arabe("Code-barres ?"), fill="black", font=font_info)

    badge_path = os.path.join(dossier_eleve, f"badge_{nom}_{prenom}.png")
    badge.save(badge_path)

    with open(FICHIER_ELEVES, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([identifiant, nom, prenom, classe, formule])

    messagebox.showinfo("Succès", f"Fiche et badge créés pour {prenom} {nom}.\nDossier créé : {dossier_eleve}")
    fen.destroy()

def supprimer_fichiers_eleve(dossier_parent, id_eleve, nom, prenom):
    dossier_eleve = os.path.join(dossier_parent, f"{nom}_{prenom}")

    try:
        if os.path.exists(dossier_eleve):
            for fichier in os.listdir(dossier_eleve):
                chemin_fichier = os.path.join(dossier_eleve, fichier)
                try:
                    if os.path.isfile(chemin_fichier):
                        os.unlink(chemin_fichier)
                except Exception as e:
                    print(f"Erreur suppression {chemin_fichier}: {e}")

            os.rmdir(dossier_eleve)
            print(f"Dossier supprimé : {dossier_eleve}")
    except Exception as e:
        print(f"Erreur suppression dossier {dossier_eleve}: {e}")

def importer_liste_eleves():
    fen_info = Toplevel()
    fen_info.transient(app)
    fen_info.grab_set()
    fen_info.title("Importation d'une liste d'élèves")
    fen_info.geometry("800x400")
    fen_info.resizable(False, False)

    fen_info.update_idletasks()
    largeur = fen_info.winfo_width()
    hauteur = fen_info.winfo_height()
    hauteur_ecran = fen_info.winfo_screenheight()
    x = (fen_info.winfo_screenwidth() // 2) - (largeur // 2)
    y = (hauteur_ecran // 2) - (hauteur // 2)
    fen_info.geometry(f"+{x}+{y}")

    message = """IMPORTANT : Avant d'importer une liste d'élèves :

1. Le fichier doit être au format Excel (.xlsx)
2. La structure du fichier doit être :
   - Colonne A : Noms des élèves
   - Colonne B : Prénoms des élèves
   - Colonne C : Classes
   - Colonne D : Formules (Demi-pension ou Externe)

3. Après l'importation, vous devrez ajouter manuellement :
   - Les photos des élèves dans leurs dossiers respectifs
     (dans le dossier fiches_eleves/NOM_PRENOM/)

Souhaitez-vous continuer avec l'importation ?
"""
    Label(fen_info, text=message, font=("Arial", 11), justify="left", wraplength=700).pack(pady=20, padx=20)

    frame_boutons = ctk.CTkFrame(fen_info, fg_color="transparent")
    frame_boutons.pack(pady=10)
    
    def annuler():
        fen_info.destroy()

    def continuer_import():
        fen_info.destroy()
        fichier_excel = filedialog.askopenfilename(
            title="Sélectionner un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
        )

        if not fichier_excel:
            return

        if not fichier_excel.lower().endswith('.xlsx'):
            messagebox.showerror("Format incorrect", "Seuls les fichiers Excel (.xlsx) sont acceptés.")
            return

        try:
            workbook = openpyxl.load_workbook(fichier_excel)
            sheet = workbook.active

            eleves_importes = []
            lignes_ignorees = 0

            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not any(row[:4]):
                    continue

                if len(row) < 4 or not all(row[:4]):
                    lignes_ignorees += 1
                    continue

                nom, prenom, classe, formule = row[:4]

                if not any(isinstance(x, str) for x in (nom, prenom, classe, formule)):
                    lignes_ignorees += 1
                    continue

                if formule.lower() not in ("demi-pension", "externe"):
                    lignes_ignorees += 1
                    continue

                eleves_importes.append((str(nom).strip(), str(prenom).strip(), str(classe).strip(), str(formule).strip()))

            if not eleves_importes:
                messagebox.showwarning("Aucun élève", "Aucun élève valide trouvé dans le fichier.")
                return

            confirmation = messagebox.askyesno(
                "Confirmation",
                f"Êtes-vous sûr de vouloir créer {len(eleves_importes)} fiches élèves ?\n"
                f"{lignes_ignorees} lignes ont été ignorées."
            )

            if not confirmation:
                return

            succes = 0
            echecs = 0

            for nom, prenom, classe, formule in eleves_importes:
                try:
                    identifiant = str(uuid.uuid4())[:8]

                    dossier_eleve = os.path.join("fiches_eleves", f"{nom}_{prenom}")
                    os.makedirs(dossier_eleve, exist_ok=True)

                    chemin_codebarre = os.path.join(dossier_eleve, f"{identifiant}.png")
                    Code128(identifiant, writer=ImageWriter()).save(chemin_codebarre[:-4])

                    nom_pdf = os.path.join(dossier_eleve, f"fiche_{nom}_{prenom}.pdf")
                    c = canvas.Canvas(nom_pdf, pagesize=A4)
                    try:
                        from reportlab.pdfbase import pdfmetrics
                        from reportlab.pdfbase.ttfonts import TTFont
                        font_path = os.path.join(os.path.dirname(__file__), "arial.ttf")
                        if os.path.exists(font_path):
                            pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                            c.setFont("ArabicFont", 20)
                            c.drawString(100, 780, preparer_texte_arabe("Fiche Élève"))
                            c.setFont("ArabicFont", 14)
                            c.drawString(100, 740, preparer_texte_arabe(f"Nom : {nom}"))
                            c.drawString(100, 720, preparer_texte_arabe(f"Prénom : {prenom}"))
                            c.drawString(100, 700, preparer_texte_arabe(f"Classe : {classe}"))
                            c.drawString(100, 680, f"Identifiant : {identifiant}")
                            c.drawString(100, 660, preparer_texte_arabe(f"Formule : {formule}"))
                        else:
                            c.setFont("Helvetica-Bold", 20)
                            c.drawString(100, 780, "Fiche Élève")
                            c.setFont("Helvetica", 14)
                            c.drawString(100, 740, f"Nom : {nom}")
                            c.drawString(100, 720, f"Prénom : {prenom}")
                            c.drawString(100, 700, f"Classe : {classe}")
                            c.drawString(100, 680, f"Identifiant : {identifiant}")
                            c.drawString(100, 660, f"Formule : {formule}")
                    except Exception as e:
                        print(f"Erreur lors de l'enregistrement de la police pour ReportLab dans l'import : {e}")
                        c.setFont("Helvetica-Bold", 20)
                        c.drawString(100, 780, "Fiche Élève")
                        c.setFont("Helvetica", 14)
                        c.drawString(100, 740, f"Nom : {nom}")
                        c.drawString(100, 720, f"Prénom : {prenom}")
                        c.drawString(100, 700, f"Classe : {classe}")
                        c.drawString(100, 680, f"Identifiant : {identifiant}")
                        c.drawString(100, 660, f"Formule : {formule}")

                    c.drawString(400, 700, preparer_texte_arabe("⚠️ Photo à ajouter manuellement"))
                    c.drawImage(chemin_codebarre, 100, 500, width=450, height=150)
                    c.save()

                    badge = Image.new("RGB", (1181, 709), "white")
                    draw = ImageDraw.Draw(badge)
                    try:
                        font_path = os.path.join(os.path.dirname(__file__), "arial.ttf")
                        if os.path.exists(font_path):
                            font_titre = ImageFont.truetype(font_path, 60)
                            font_info = ImageFont.truetype(font_path, 40)
                        else:
                            font_titre = ImageFont.load_default()
                            font_info = ImageFont.load_default()
                    except Exception as e:
                        print(f"Erreur lors du chargement de la police pour PIL dans l'import : {e}")
                        font_titre = ImageFont.load_default()
                        font_info = ImageFont.load_default()

                    draw.text((480, 20), preparer_texte_arabe("Carte d'accès"), fill="black", font=font_titre)
                    draw.text((40, 150), preparer_texte_arabe(f"Nom : {nom.upper()}"), fill="black", font=font_info)
                    draw.text((40, 220), preparer_texte_arabe(f"Prénom : {prenom.upper()}"), fill="black", font=font_info)
                    draw.text((40, 290), preparer_texte_arabe(f"Classe : {classe}"), fill="black", font=font_info)
                    draw.text((40, 360), f"ID : {identifiant}", fill="black", font=font_info)
                    draw.text((40, 430), preparer_texte_arabe(f"Formule : {formule}"), fill="black", font=font_info)
                    draw.text((850, 150), preparer_texte_arabe("⚠️ Photo à ajouter"), fill="black", font=font_info)

                    if os.path.exists(chemin_codebarre):
                        code_img = Image.open(chemin_codebarre).resize((750, 225))
                        badge.paste(code_img, (215, 490))
                    else:
                        draw.text((540, 600), preparer_texte_arabe("Code-barres ?"), fill="black", font=font_info)

                    badge_path = os.path.join(dossier_eleve, f"badge_{nom}_{prenom}.png")
                    badge.save(badge_path)

                    with open(FICHIER_ELEVES, "a", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerow([identifiant, nom, prenom, classe, formule])

                    succes += 1

                except Exception as e:
                    print(f"Erreur création fiche pour {nom} {prenom}: {str(e)}")
                    echecs += 1

            messagebox.showinfo(
                "Importation terminée",
                f"Résultat de l'importation :\n\n"
                f"- Fiches créées avec succès : {succes}\n"
                f"- Échecs de création : {echecs}\n\n"
                f"Note : Vous devez ajouter manuellement les photos des élèves\n"
                f"dans leurs dossiers respectifs."
            )

        except Exception as e:
            messagebox.showerror(
                "Erreur d'importation",
                f"Une erreur est survenue lors de la lecture du fichier Excel :\n\n{str(e)}"
            )

    ctk.CTkButton(frame_boutons, text="Annuler", width=130, height=35, font=("Arial", 16),
              command=annuler).pack(side="left", padx=10)

    ctk.CTkButton(frame_boutons, text="Importer", width=130, height=35, font=("Arial", 16),
              command=continuer_import).pack(side="left", padx=10)

def ouvrir_fenetre_scan():
    fen = tk.Toplevel()
    fen.title("Contrôle Cantine")
    fen.state('zoomed')

    ctk.CTkLabel(fen, text="Scanner / Saisir ID Code-Barres :", font=("Arial", 16)).pack(pady=(15, 1))

    entry_id = ctk.CTkEntry(fen, font=("Arial", 17), width=350, height=40, placeholder_text="ID code-barres")
    entry_id.pack(pady=5)
    entry_id.focus()

    colonnes = ("ID", "Nom", "Prénom", "Classe", "Date", "Heure")
    tableau = ttk.Treeview(fen, columns=colonnes, show="headings", height=20)
    for col in colonnes:
        tableau.heading(col, text=col)
        tableau.column(col, anchor="center", width=120)

    scrollbar = Scrollbar(fen, orient="vertical", command=tableau.yview)
    tableau.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    tableau.pack(pady=(5), fill="both", expand=True)

    def ajouter_ligne(data):
        tableau.insert("", "end", values=data)

    def charger_ids_du_jour(fichier):
        ids = set()
        if os.path.exists(fichier):
            with open(fichier, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for ligne in reader:
                    ids.add(ligne["ID"])
        return ids

    def charger_eleves():
        eleves = {}
        with open(FICHIER_ELEVES, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for ligne in reader:
                eleves[ligne["ID"]] = (ligne["Nom"], ligne["Prénom"], ligne["Classe"])
        return eleves

    def traiter_scan(event=None):
        identifiant = entry_id.get().strip()
        if not identifiant:
            return

        maintenant = datetime.now()
        date_str = maintenant.strftime("%Y-%m-%d")
        heure_str = maintenant.strftime("%H:%M:%S")

        dossier_jour = "passages"
        os.makedirs(dossier_jour, exist_ok=True)
        fichier_jour = os.path.join(dossier_jour, f"{date_str}.csv")

        ids_jour = charger_ids_du_jour(fichier_jour)

        if identifiant in ids_jour:
            jouer_son("error.wav")
            messagebox.showwarning("Déjà scanné", "Cet élève est déjà passé aujourd'hui.", parent=fen)
            entry_id.delete(0, END)
            fen.after(100, lambda: entry_id.focus_force())
            return

        eleves = {}
        formule_eleve = None
        with open(FICHIER_ELEVES, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for ligne in reader:
                if ligne["ID"] == identifiant:
                    eleves[identifiant] = (ligne["Nom"], ligne["Prénom"], ligne["Classe"])
                    if "Formule" in ligne:
                        formule_eleve = ligne["Formule"]
                    break

        if identifiant not in eleves:
            jouer_son("error.wav")
            messagebox.showerror("Inconnu", "Cet identifiant n'existe pas dans la base de données.", parent=fen)
            entry_id.delete(0, END)
            fen.after(100, lambda: entry_id.focus_force())
            return

        if formule_eleve and formule_eleve.lower() == "externe":
            jouer_son("error.wav")
            messagebox.showwarning("Accès refusé",
                                "Cet élève n'a pas le droit d'accès à la cantine\n"
                                f"Raison : Formule '{formule_eleve}'", parent=fen)
            entry_id.delete(0, END)
            fen.after(100, lambda: entry_id.focus_force())
            return

        nom, prenom, classe = eleves[identifiant]

        ligne = [identifiant, nom, prenom, classe, date_str, heure_str]

        existe = os.path.exists(fichier_jour)
        with open(fichier_jour, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not existe:
                writer.writerow(["ID", "Nom", "Prénom", "Classe", "Date", "Heure"])
            writer.writerow(ligne)

        with open(FICHIER_PRESENCES, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(ligne)

        jouer_son("success.wav")
        ajouter_ligne(ligne)
        entry_id.delete(0, END)

    entry_id.bind("<Return>", traiter_scan)
    ctk.CTkButton(fen, text="Valider l'entrée", font=("Arial", 14), width=200, height=40, command=traiter_scan).pack(pady=10)

    maintenant = datetime.now()
    date_str = maintenant.strftime("%Y-%m-%d")
    fichier_jour = os.path.join("passages", f"{date_str}.csv")
    if os.path.exists(fichier_jour):
        with open(fichier_jour, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for ligne in reader:
                ajouter_ligne((ligne["ID"], ligne["Nom"], ligne["Prénom"], ligne["Classe"], ligne["Date"], ligne["Heure"]))

def generer_pdf_badges(eleves_selectionnes, dossier_parent="fiches_eleves"):
    try:
        if not eleves_selectionnes:
            messagebox.showwarning("Aucun élève", "Aucun élève sélectionné pour l'impression.")
            return False

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(False)

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        nom_fichier = f"badges_eleves_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        chemin_complet = os.path.join(desktop, nom_fichier)

        largeur_badge = 90
        hauteur_badge = 70
        badges_par_page = 8
        badges_par_ligne = 2
        lignes_par_page = 4

        marge_gauche = (210 - (badges_par_ligne * largeur_badge)) / 2
        marge_haut = 6
        espace_entre_badges = 0
        espace_entre_lignes = 0

        epaisseur_contour = 0.3
        decalage_contour = epaisseur_contour / 2

        compteur = 0

        for eleve in eleves_selectionnes:
            id_eleve = eleve[0]
            nom_eleve = eleve[1]
            prenom_eleve = eleve[2]

            dossier_eleve = os.path.join(dossier_parent, f"{nom_eleve}_{prenom_eleve}")
            badge_path = os.path.join(dossier_eleve, f"badge_{nom_eleve}_{prenom_eleve}.png")

            if not os.path.exists(badge_path):
                print(f"Fichier badge introuvable: {badge_path}")
                continue

            if compteur % badges_par_page == 0:
                pdf.add_page()
                compteur = 0

            ligne = (compteur // badges_par_ligne) % lignes_par_page
            colonne = compteur % badges_par_ligne

            x = marge_gauche + (colonne * (largeur_badge + espace_entre_badges))
            y = marge_haut + (ligne * (hauteur_badge + espace_entre_lignes))

            pdf.set_draw_color(0, 0, 0)
            pdf.set_line_width(epaisseur_contour)
            pdf.rect(x - decalage_contour,
                    y - decalage_contour,
                    largeur_badge + epaisseur_contour,
                    hauteur_badge + epaisseur_contour)

            pdf.image(badge_path,
                     x=x,
                     y=y,
                     w=largeur_badge,
                     h=hauteur_badge)

            compteur += 1

        if compteur == 0:
            messagebox.showerror("Erreur", "Aucun badge trouvé pour les élèves sélectionnés.")
            return False

        pdf.output(chemin_complet)

        if os.name == 'nt':
            os.startfile(chemin_complet)
        elif os.name == 'posix':
            subprocess.run(['open', chemin_complet] if sys.platform == 'darwin' else ['xdg-open', chemin_complet])

        messagebox.showinfo("Succès", f"PDF généré avec 8 badges par page (9x7cm):\n{chemin_complet}")
        return True

    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de la génération:\n{str(e)}")
        return False

def supprimer_toutes_donnees():
    fen_confirm = Toplevel()
    fen_confirm.title("Confirmation de suppression")
    fen_confirm.geometry("600x300")
    fen_confirm.resizable(False, False)
    fen_confirm.transient(app)
    fen_confirm.grab_set()

    fen_confirm.update_idletasks()
    largeur = fen_confirm.winfo_width()
    hauteur = fen_confirm.winfo_height()
    hauteur_ecran = fen_confirm.winfo_screenheight()
    x = (fen_confirm.winfo_screenwidth() // 2) - (largeur // 2)
    y = (hauteur_ecran // 2) - (hauteur // 2)
    fen_confirm.geometry(f"+{x}+{y}")

    message = """ATTENTION : Suppression de toutes les données

Cette action va supprimer toutes les données :
   - Tous les élèves et leurs dossiers
   - Tous les badges
   - Tous les rapports de présence
   - Tous les fichiers de passages quotidiens

Souhaitez-vous vraiment continuer ?
"""
    Label(fen_confirm, text=message, font=("Arial", 11),
          justify="left", wraplength=550, fg="red").pack(pady=20, padx=20)

    frame_boutons = ctk.CTkFrame(fen_confirm, fg_color="transparent")
    frame_boutons.pack(pady=20)

    def annuler():
        fen_confirm.destroy()

    def executer_suppression():
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dossier_backup = os.path.join("backup & restore", f"backup_{timestamp}")
            os.makedirs(dossier_backup, exist_ok=True)

            elements_a_sauvegarder = [
                "fiches_eleves",
                "liste_eleves.csv",
                "presences_cantine.csv",
                "passages"
            ]

            for element in elements_a_sauvegarder:
                try:
                    if os.path.isdir(element):
                        shutil.copytree(element, os.path.join(dossier_backup, element))
                    elif os.path.isfile(element):
                        shutil.copy2(element, dossier_backup)
                except Exception as e:
                    print(f"Erreur sauvegarde {element}: {e}")

            try:
                if os.path.exists("fiches_eleves"):
                    shutil.rmtree("fiches_eleves")

                if os.path.exists("passages"):
                    shutil.rmtree("passages")

                for fichier in [FICHIER_ELEVES, FICHIER_PRESENCES]:
                    with open(fichier, "w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        if fichier == FICHIER_ELEVES:
                            writer.writerow(["ID", "Nom", "Prénom", "Classe", "Formule"])
                        else:
                            writer.writerow(["ID", "Nom", "Prénom", "Classe", "Date", "Heure"])

                messagebox.showinfo(
                    "Succès",
                    f"Toutes les données ont été supprimées.\n"
                    f"Une sauvegarde a été créée dans :\n{dossier_backup}"
                )

            except Exception as e:
                messagebox.showerror(
                    "Erreur de suppression",
                    f"Une erreur est survenue lors de la suppression :\n{str(e)}\n"
                    f"Les données sauvegardées sont disponibles dans :\n{dossier_backup}"
                )

        except Exception as e:
            messagebox.showerror(
                "Erreur critique",
                f"Une erreur est survenue :\n{str(e)}\n"
                "Aucune donnée n'a été modifiée."
            )

        fen_confirm.destroy()
        
            # --- Boutons de la fenêtre de confirmation ---
    ctk.CTkButton(frame_boutons, text="Annuler", width=150, font=("Arial", 13),
              command=fen_confirm.destroy).pack(side="left", padx=10)

    ctk.CTkButton(frame_boutons, text="Confirmer", width=150, font=("Arial", 13),
              fg_color="#ffcccc", hover_color="#ff9999", text_color="black",
              command=executer_suppression).pack(side="left", padx=10)
    # --- Fin des boutons ---


def ouvrir_fenetre_rapport():
    from tkcalendar import DateEntry
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from tkinter import filedialog
    import os

    fen = tk.Toplevel()
    fen.title("Rapport des passages")
    fen.state('zoomed')

    Label(fen, text="Sélectionner une période :", font=("Arial", 12)).pack(pady=5)

    cadre_dates = Frame(fen)
    cadre_dates.pack(pady=5)

    Label(cadre_dates, text="Date de début:", font=("Arial", 11)).grid(row=0, column=0, padx=10)
    date_debut = DateEntry(cadre_dates, width=12, background='darkblue', foreground='white',
                           borderwidth=2, date_pattern='yyyy-mm-dd')
    date_debut.grid(row=0, column=1)

    Label(cadre_dates, text="Date de fin:", font=("Arial", 11)).grid(row=0, column=2, padx=10)
    date_fin = DateEntry(cadre_dates, width=12, background='darkblue', foreground='white',
                         borderwidth=2, date_pattern='yyyy-mm-dd')
    date_fin.grid(row=0, column=3)

    colonnes = ("ID", "Nom", "Prénom", "Classe", "Date", "Heure")
    tableau = ttk.Treeview(fen, columns=colonnes, show="headings", height=20)
    for col in colonnes:
        tableau.heading(col, text=col)
        tableau.column(col, anchor="center", width=120)
    tableau.pack(pady=10, fill="both", expand=True)

    label_total = Label(fen, text="", font=("Arial", 12, "bold"))
    label_total.pack(pady=5)

    def charger_rapport():
        tableau.delete(*tableau.get_children())
        total = 0
        date_d = datetime.strptime(date_debut.get(), "%Y-%m-%d").date()
        date_f = datetime.strptime(date_fin.get(), "%Y-%m-%d").date()

        with open(FICHIER_PRESENCES, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for ligne in reader:
                try:
                    d = datetime.strptime(ligne["Date"], "%Y-%m-%d").date()
                    if date_d <= d <= date_f:
                        tableau.insert("", "end", values=(
                            ligne["ID"], ligne["Nom"], ligne["Prénom"],
                            ligne["Classe"], ligne["Date"], ligne["Heure"]
                        ))
                        total += 1
                except ValueError:
                    continue

        label_total.config(text=f"Nombre total de passages : {total}")

    def exporter_pdf():
        elements = tableau.get_children()
        if not elements:
            messagebox.showinfo("Info", "Aucun rapport à exporter.", parent=fen)
            return

        date_d = date_debut.get()
        date_f = date_fin.get()
        nom_fichier = f"rapport_{date_d}_au_{date_f}.pdf"

        dossier = filedialog.askdirectory(title="Choisir le dossier de destination")
        if not dossier:
            return

        chemin_fichier = os.path.join(dossier, nom_fichier)

        c = canvas.Canvas(chemin_fichier, pagesize=A4)
        largeur, hauteur = A4
        y = hauteur - 50

        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            font_path = os.path.join(os.path.dirname(__file__), "arial.ttf")
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                font_name_bold = 'ArabicFont'
                font_name_normal = 'ArabicFont'
            else:
                messagebox.showwarning("Police manquante", "Le fichier 'arial.ttf' est introuvable. Le texte arabe pourrait ne pas s'afficher correctement dans le rapport PDF.")
                font_name_bold = "Helvetica-Bold"
                font_name_normal = "Helvetica"
        except Exception as e:
            print(f"Erreur lors de l'enregistrement de la police pour ReportLab (rapport): {e}")
            font_name_bold = "Helvetica-Bold"
            font_name_normal = "Helvetica"

        c.setFont(font_name_bold, 14)
        c.drawString(50, y, preparer_texte_arabe(f"Rapport des passages du {date_d} au {date_f}"))
        y -= 30

        c.setFont(font_name_bold, 10)
        entetes = ["ID", "Nom", "Prénom", "Classe", "Date", "Heure"]
        x_offsets = [50, 130, 210, 290, 370, 450]
        for i, en in enumerate(entetes):
            c.drawString(x_offsets[i], y, preparer_texte_arabe(en))
        y -= 20

        c.setFont(font_name_normal, 10)
        total = 0
        for ligne in elements:
            if y < 50:
                c.showPage()
                y = hauteur - 50
                c.setFont(font_name_normal, 10)
            valeurs = tableau.item(ligne)["values"]

            display_values = [
                str(valeurs[0]),
                preparer_texte_arabe(str(valeurs[1])),
                preparer_texte_arabe(str(valeurs[2])),
                preparer_texte_arabe(str(valeurs[3])),
                str(valeurs[4]),
                str(valeurs[5])
            ]

            for i, val in enumerate(display_values):
                c.drawString(x_offsets[i], y, val)
            y -= 20
            total += 1

        if y < 70:
            c.showPage()
            y = hauteur - 50
        y -= 10
        c.setFont(font_name_bold, 12)
        c.drawString(50, y, preparer_texte_arabe(f"Nombre total de passages : {total}"))

        c.save()
        messagebox.showinfo("Exportation réussie", f"Le rapport a été enregistré dans :\n{chemin_fichier}", parent=fen)

    button_frame = ctk.CTkFrame(fen, fg_color="transparent")
    button_frame.pack(pady=(10, 20))

    ctk.CTkButton(button_frame, text="Afficher Rapport", font=("Arial", 15), width=170, height=40,
                  command=charger_rapport).pack(side="left", padx=10)

    ctk.CTkButton(button_frame, text="Exporter en PDF", font=("Arial", 15), width=170, height=40,
                  command=exporter_pdf).pack(side="left", padx=10)

    ctk.CTkButton(button_frame, text="Fermer", font=("Arial", 15), width=130, height=40,
                  command=fen.destroy).pack(side="left", padx=10)

def ouvrir_liste_eleves():
    
    fen = tk.Toplevel()
    fen.state('zoomed')
    fen.title("Liste des élèves inscrits")
    fen.grab_set()

    

    top_frame = Frame(fen)
    top_frame.pack(fill="x", pady=10, padx=20)

    ctk.CTkLabel(top_frame, text="Rechercher par nom :", font=("Arial", 13)).pack(side="left", padx=(5, 10))
    champ_recherche = ctk.CTkEntry(top_frame, font=("Arial", 13), width=250, placeholder_text="Tapez un nom ou prénom")
    champ_recherche.pack(side="left", padx=(5, 5))

    total_var = StringVar()
    demi_pension_var = StringVar()
    externe_var = StringVar()

    total_var.set("Total élèves : 0")
    demi_pension_var.set("Demi-pension : 0")
    externe_var.set("Externe : 0")
    
    label_demi = Label(top_frame, textvariable=demi_pension_var, font=("Arial", 12, "bold"), fg="black")
    label_demi.pack(side="right", padx=(10, 50))

    label_externe = Label(top_frame, textvariable=externe_var, font=("Arial", 12, "bold"), fg="black")
    label_externe.pack(side="right", padx=(10, 30))
    
    label_total = Label(top_frame, textvariable=total_var, font=("Arial", 12, "bold"), fg="black")
    label_total.pack(side="right", padx=(10, 30))

    colonnes = ("ID", "Nom", "Prénom", "Classe", "Formule")
    tableau = ttk.Treeview(fen, columns=colonnes, show="headings", selectmode="extended")

    tableau.column("ID", width=100, anchor="center")
    tableau.column("Nom", width=150, anchor="center")
    tableau.column("Prénom", width=150, anchor="center")
    tableau.column("Classe", width=150, anchor="center")
    tableau.column("Formule", width=150, anchor="center")

    for col in colonnes:
        tableau.heading(col, text=col)

    scrollbar = Scrollbar(fen, orient="vertical", command=tableau.yview)
    tableau.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    tableau.pack(expand=True, fill="both", padx=10, pady=5)

    def mettre_a_jour_total():
        total = 0
        demi = 0
        externe = 0

        for item in tableau.get_children():
            total += 1
            valeurs = tableau.item(item)["values"]
            if len(valeurs) >= 5:
                formule = valeurs[4].strip().lower()
                if "demi" in formule:
                    demi += 1
                elif "externe" in formule:
                    externe += 1

        total_var.set(f"Total élèves : {total}")
        demi_pension_var.set(f"Demi-pension : {demi}")
        externe_var.set(f"Externe : {externe}")


    def rechercher_nom(event=None):
        query = champ_recherche.get().strip().lower()
        tableau.delete(*tableau.get_children())
        with open(FICHIER_ELEVES, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for index, row in enumerate(reader):
                if index == 0 and row[0].lower() in ["id", "identifiant"]:
                    continue
                if len(row) > 1 and query in row[1].lower():
                    tableau.insert("", "end", values=row)
                elif len(row) > 2 and query in row[2].lower():
                    tableau.insert("", "end", values=row)
        mettre_a_jour_total()

    champ_recherche.bind("<KeyRelease>", rechercher_nom)

    def charger_eleves():
        tableau.delete(*tableau.get_children())
        try:
            with open(FICHIER_ELEVES, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header and len(header) < 5:
                    messagebox.showwarning("Fichier CSV obsolète", "Le fichier 'liste_eleves.csv' semble ne pas avoir la colonne 'Formule'. Veuillez recréer les fiches ou mettre à jour le fichier manuellement si nécessaire.")
                    with open(FICHIER_ELEVES, "w", newline="", encoding="utf-8") as fw:
                        writer = csv.writer(fw)
                        writer.writerow(["ID", "Nom", "Prénom", "Classe", "Formule"])
                    f.seek(0)
                    next(reader, None)

                for row in reader:
                    if len(row) < 5:
                        row.append("N/A")
                    tableau.insert("", "end", values=row)
        except Exception as e:
            messagebox.showerror("Erreur de lecture CSV", f"Erreur lors du chargement des élèves : {e}")
        mettre_a_jour_total()

    def imprimer_badges():
        selections = tableau.selection()
        if not selections:
            messagebox.showwarning("Aucune sélection", "Veuillez sélectionner au moins un élève.", parent=fen)
            return

        eleves_selectionnes = []
        for item in selections:
            valeurs = tableau.item(item)["values"]
            eleves_selectionnes.append(valeurs)

        succes = generer_pdf_badges(eleves_selectionnes)
        if succes:
            messagebox.showinfo("Succès", "Les badges ont été générés avec succès et sauvegardés sur le Bureau.", parent=fen)

    def exporter_pdf():
        lignes = tableau.get_children()
        if not lignes:
            messagebox.showwarning("Aucun élève", "Aucune donnée à exporter.", parent=fen)
            return

        pdf = FPDF()
        pdf.add_page()

        font_path = os.path.join(os.path.dirname(__file__), "arial.ttf")
        font_path_bold = os.path.join(os.path.dirname(__file__), "arialbd.ttf")

        if not os.path.exists(font_path):
            messagebox.showwarning("Police manquante", "Le fichier 'arial.ttf' est introuvable. Le texte arabe pourrait ne pas s'afficher correctement dans le PDF exporté.")
            pdf.set_font("Helvetica", "", 12)
        else:
            pdf.add_font("DejaVuSans", "", font_path, uni=True)
            if os.path.exists(font_path_bold):
                pdf.add_font("DejaVuSans", "B", font_path_bold, uni=True)
            else:
                messagebox.showwarning("Police Bold manquante", "Le fichier 'arialbd.ttf' pour la police grasse est introuvable. La mise en gras pourrait ne pas être exacte dans le PDF exporté.")
                pdf.add_font("DejaVuSans", "B", font_path, uni=True)

            pdf.set_font("DejaVuSans", "B", 14)
            pdf.cell(0, 10, preparer_texte_arabe("Liste des élèves inscrits"), ln=True, align="C")
            pdf.ln(10)

            widths = {
                'ID': 30,
                'Nom': 45,
                'Prénom': 45,
                'Classe': 40,
                'Formule': 30
            }

            pdf.set_font("DejaVuSans", "B", 12)
            pdf.cell(widths['ID'], 10, preparer_texte_arabe("ID"), 1)
            pdf.cell(widths['Nom'], 10, preparer_texte_arabe("Nom"), 1)
            pdf.cell(widths['Prénom'], 10, preparer_texte_arabe("Prénom"), 1)
            pdf.cell(widths['Classe'], 10, preparer_texte_arabe("Classe"), 1)
            pdf.cell(widths['Formule'], 10, preparer_texte_arabe("Formule"), 1)
            pdf.ln()

            pdf.set_font("DejaVuSans", "", 12)
            for item in lignes:
                valeurs = tableau.item(item)["values"]

                display_values = [
                    str(valeurs[0]),
                    preparer_texte_arabe(str(valeurs[1])),
                    preparer_texte_arabe(str(valeurs[2])),
                    preparer_texte_arabe(str(valeurs[3])),
                    preparer_texte_arabe(str(valeurs[4]) if len(valeurs) > 4 else "N/A")
                ]

                pdf.cell(widths['ID'], 10, display_values[0], 1)
                pdf.cell(widths['Nom'], 10, display_values[1], 1)
                pdf.cell(widths['Prénom'], 10, display_values[2], 1)
                pdf.cell(widths['Classe'], 10, display_values[3], 1)
                pdf.cell(widths['Formule'], 10, display_values[4], 1)
                pdf.ln()

        try:
            nom_fichier = f"liste_eleves_{datetime.now().strftime('%Y-%m-%d')}.pdf"
            chemin_bureau = os.path.join(os.path.expanduser("~"), "Desktop", nom_fichier)
            pdf.output(chemin_bureau)
            messagebox.showinfo("Export PDF", f"PDF exporté sur le Bureau :\n{chemin_bureau}", parent=fen)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export : {e}", parent=fen)

    def supprimer_eleve():
        selections = tableau.selection()
        if not selections:
            messagebox.showwarning("Aucune sélection", "Veuillez sélectionner au moins un élève à supprimer.", parent=fen)
            return

        eleves_a_supprimer = []
        for item in selections:
            valeurs = tableau.item(item)["values"]
            eleves_a_supprimer.append({
                'id': valeurs[0],
                'nom': valeurs[1],
                'prenom': valeurs[2]
            })

        confirmation = messagebox.askyesno(
            "Confirmer la suppression",
            f"Supprimer {len(eleves_a_supprimer)} élève(s) sélectionné(s) ?\n"
            "Cette action supprimera aussi leurs fiches et badges.", parent=fen
        )
        if not confirmation:
            return

        for eleve in eleves_a_supprimer:
            supprimer_fichiers_eleve("fiches_eleves", eleve['id'], eleve['nom'], eleve['prenom'])

        nouvelles_lignes = []
        try:
            with open(FICHIER_ELEVES, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader)
                nouvelles_lignes.append(header)

                for ligne in reader:
                    if not any(ligne[0] == eleve['id'] for eleve in eleves_a_supprimer):
                        nouvelles_lignes.append(ligne)

            with open(FICHIER_ELEVES, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(nouvelles_lignes)

            charger_eleves()
            messagebox.showinfo("Suppression réussie",
                              f"{len(eleves_a_supprimer)} élève(s) supprimé(s)\n"
                              "avec leurs fiches et badges associés.", parent=fen)
        except Exception as e:
            messagebox.showerror("Erreur de suppression", f"Une erreur est survenue lors de la suppression des données du fichier CSV : {e}")

    frame_boutons = ctk.CTkFrame(fen, fg_color="transparent")
    frame_boutons.pack(pady=(10, 20))

    ctk.CTkButton(frame_boutons, text="Exporter en PDF", font=("Arial", 15), width=170, height=40,
              command=exporter_pdf).pack(side="left", padx=10)

    ctk.CTkButton(frame_boutons, text="Imprimer badge(s)", font=("Arial", 15), width=170, height=40,
              command=imprimer_badges).pack(side="left", padx=10)

    ctk.CTkButton(frame_boutons, text="Supprimer", font=("Arial", 15), width=130, height=40,
              fg_color="#ffcccc", hover_color="#ff9999", text_color="black",
              command=supprimer_eleve).pack(side="left", padx=10)

    ctk.CTkButton(frame_boutons, text="Fermer", font=("Arial", 15), width=130, height=40,
              command=fen.destroy).pack(side="left", padx=10)

    # L'appel manquant pour charger les élèves au démarrage
    charger_eleves()
    
def lancer_application():
    global app
    app = tk.Tk()
    app.title("Menu Administrateur")
    app.state('zoomed')

    menu_bar = Menu(app)
    menu_aide = Menu(menu_bar, tearoff=0)
    menu_aide.add_command(label="Version", command=lambda: messagebox.showinfo("Version", "Version 1.0\nCréé par Aoudjit kosseila"))
    menu_aide.add_command(label="Aide", command=lambda: messagebox.showinfo("Aide", "Pour toute assistance, contactez le développeur du logiciel."))
    menu_bar.add_cascade(label="Aide", menu=menu_aide)
    app.config(menu=menu_bar)

    ctk.CTkLabel(app, text="Menu Administrateur", font=("Arial", 20, "bold")).pack(pady=40)

    ctk.CTkButton(app, text="Créer une fiche élève", width=300, height=40, font=("Arial", 16),
                  command=ouvrir_fenetre_fiche).pack(pady=10)
    ctk.CTkButton(app, text="Importer une liste d'élèves", width=300, height=40, font=("Arial", 16),
                  command=importer_liste_eleves).pack(pady=10)
    ctk.CTkButton(app, text="Liste des élèves inscrits", width=300, height=40, font=("Arial", 16),
                  command=ouvrir_liste_eleves).pack(pady=10)
    ctk.CTkButton(app, text="Contrôle d'accès", width=300, height=40, font=("Arial", 16),
                  command=ouvrir_fenetre_scan).pack(pady=10)
    ctk.CTkButton(app, text="Exporter un rapport", width=300, height=40, font=("Arial", 16),
                  command=ouvrir_fenetre_rapport).pack(pady=10)
    ctk.CTkButton(app, text="Supprimer toutes les données", width=300, height=40, font=("Arial", 16),
                  fg_color="#ffcccc", hover_color="#ff9999", text_color="black",
                  command=supprimer_toutes_donnees).pack(pady=10)
    ctk.CTkButton(app, text="Quitter", width=300, height=40, font=("Arial", 16),
                  command=app.destroy).pack(pady=10)

    app.mainloop()

admin_icon_ctk = None
controle_icon = None
logo_img = None
background_image_tk = None
racine = None # Define racine globally or pass it as an argument


def jouer_son(chemin_fichier):
    try:
        son = pygame.mixer.Sound(chemin_fichier)
        son.set_volume(1.0)  # Volume max
        son.play()
    except Exception as e:
        print(f"Erreur lecture son {chemin_fichier} : {e}")
        
        

def ouvrir_fenetre_version():
    statut = "Activée ✅" if verifier_licence_locale() else "Non activée ❌"

    fenetre_version = Toplevel()
    fenetre_version.title("Version et activation")
    fenetre_version.geometry("400x200")
    fenetre_version.resizable(False, False)

    Label(fenetre_version, text="CantiTrack - Version 1.0", font=("Arial", 14, "bold")).pack(pady=10)
    Label(fenetre_version, text=f"Statut d'activation : {statut}", font=("Arial", 12)).pack(pady=10)

    if not verifier_licence_locale():
        Button(fenetre_version, text="Activer maintenant", command=lambda: demander_activation(fenetre_version)).pack(pady=15)

    Button(fenetre_version, text="Fermer", command=fenetre_version.destroy).pack(pady=10)


        

def ouvrir_choix_utilisateur():
    global racine, logo_img, admin_icon_ctk, acces_icon_ctk, background_image_tk, admin_button_reference

    racine = tk.Tk()
    racine.title("CantiTrack - Sélection du rôle")
    racine.state('zoomed')

    try:
        dossier_script = os.path.dirname(os.path.abspath(__file__))
        chemin_fond = os.path.join(dossier_script, "fond.jpg")

        background_image = Image.open(chemin_fond)
        background_image_tk = ImageTk.PhotoImage(background_image)

        canvas = tk.Canvas(racine, highlightthickness=0)
        canvas.pack(fill="both", expand=True)

        def resize_background(event):
            global background_image_tk
            new_width = event.width
            new_height = event.height

            if new_width <= 0 or new_height <= 0:
                return

            resized_image = background_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
            background_image_tk = ImageTk.PhotoImage(resized_image)
            canvas.delete("background")
            canvas.create_image(0, 0, image=background_image_tk, anchor="nw", tags="background")
            canvas.lower("background")

        canvas.bind("<Configure>", resize_background)

        racine.update_idletasks()
        initial_width = racine.winfo_width()
        initial_height = racine.winfo_height()
        if initial_width > 0 and initial_height > 0:
            resized_image_initial = background_image.resize((initial_width, initial_height), Image.Resampling.LANCZOS)
            background_image_tk = ImageTk.PhotoImage(resized_image_initial)
            canvas.create_image(0, 0, image=background_image_tk, anchor="nw", tags="background")
            canvas.lower("background")
        else:
            canvas.create_image(0, 0, image=background_image_tk, anchor="nw", tags="background")
            canvas.lower("background")

    except Exception as e:
        print(f"Erreur de chargement de l'image de fond: {e}")
        racine.configure(bg="#FFFFFF")
        canvas = tk.Canvas(racine, bg="#FFFFFF", highlightthickness=0)
        canvas.pack(fill="both", expand=True)

    menu_bar = Menu(racine)
    menu_aide = Menu(menu_bar, tearoff=0)
    menu_aide.add_command(label="Version", command=ouvrir_fenetre_version)
    menu_aide.add_command(label="Aide", command=lambda: messagebox.showinfo("Aide", "Pour toute assistance, contactez le développeur."))
    menu_aide.add_command(label="Mise à jour", command=verifier_mise_a_jour)
    menu_bar.add_cascade(label="Aide", menu=menu_aide)
    racine.config(menu=menu_bar)

    def update_element_positions(event=None):
        global admin_button_reference
        canvas_width = canvas.winfo_width()
        canvas_height = canvas.winfo_height()

        if canvas_width <= 0 or canvas_height <= 0:
            return

        logo_height = 150
        logo_pady_bottom = 20
        cantitrack_text_height = 40
        cantitrack_pady_bottom = 10
        subtitle_text_height = 20
        subtitle_pady_bottom = 40
        choix_role_text_height = 20
        choix_role_pady_bottom = 5
        button_height = 40
        button_pady = 10
        button_width = 300

        total_content_height = (logo_height + logo_pady_bottom +
                                cantitrack_text_height + cantitrack_pady_bottom +
                                subtitle_text_height + subtitle_pady_bottom +
                                choix_role_text_height + choix_role_pady_bottom +
                                button_height + button_pady +
                                button_height + button_pady)

        start_y = (canvas_height - total_content_height) / 2
        if start_y < 0:
            start_y = 0

        current_y = start_y
        center_x = canvas_width / 2

        cantitrack_x = center_x + (cantitrack_text_height / 0.11)
        cantitrack_y = current_y + (cantitrack_text_height / 2)
        if not hasattr(update_element_positions, 'cantitrack_text_item_id'):
            update_element_positions.cantitrack_text_item_id = canvas.create_text(cantitrack_x, cantitrack_y, text="Bienvenue sur CantiTrack", font=("Arial", 30, "bold"), fill="#2c3e50", tags="cantitrack_text")
        else:
            canvas.itemconfig(update_element_positions.cantitrack_text_item_id, text="Bienvenue sur CantiTrack")
            canvas.coords(update_element_positions.cantitrack_text_item_id, cantitrack_x, cantitrack_y)
        current_y += cantitrack_text_height + cantitrack_pady_bottom

        subtitle_x = center_x + (subtitle_text_height / 0.055)
        subtitle_y = current_y + (subtitle_text_height / 2)
        if not hasattr(update_element_positions, 'subtitle_text_item_id'):
            update_element_positions.subtitle_text_item_id = canvas.create_text(
                subtitle_x, subtitle_y,
                text="Votre solution complète pour une gestion simple et efficace de la cantine scolaire.",
                font=("Arial", 11), fill="#555555", tags="subtitle_text")
        else:
            canvas.coords(update_element_positions.subtitle_text_item_id, subtitle_x, subtitle_y)
        current_y += subtitle_text_height + subtitle_pady_bottom

        choix_role_x = center_x + (choix_role_text_height / 0.057)
        choix_role_y = current_y + (choix_role_text_height / 0.4)
        if not hasattr(update_element_positions, 'choix_role_text_item_id'):
            update_element_positions.choix_role_text_item_id = canvas.create_text(choix_role_x, choix_role_y, text="Choisissez votre rôle", font=("Arial", 13), fill="black", tags="choix_role_text")
        else:
            canvas.coords(update_element_positions.choix_role_text_item_id, choix_role_x, choix_role_y)
        current_y += choix_role_text_height + choix_role_pady_bottom

        admin_btn_x = center_x + (button_height / 0.111)
        admin_btn_y = current_y + (button_height / 0.5)
        if not hasattr(update_element_positions, 'admin_button_window_id'):
            global admin_icon_ctk
            try:
                chemin_admin_icon = os.path.join(dossier_script, "Administrateur.png")
                admin_icon_ctk = ctk.CTkImage(light_image=Image.open(chemin_admin_icon), dark_image=Image.open(chemin_admin_icon), size=(25, 25))
            except Exception as e:
                print(f"Erreur de chargement de l'icône Administrateur.png: {e}")
                admin_icon_ctk = None

            if verifier_licence_locale():
                admin_button = ctk.CTkButton(canvas, text="Administrateur", width=button_width, height=button_height, font=("Arial", 14, "bold"), image=admin_icon_ctk, compound="left", fg_color="#FFFFFF", hover_color="#E6E6FF", text_color="#4C4C4C", border_color="#808080", border_width=1, command=lambda: verifier_mot_de_passe())
            else:
                def afficher_non_active():
                    messagebox.showerror("Licence requise", "Veuillez activer votre application pour accéder à cet espace.")
                admin_button = ctk.CTkButton(canvas, text="Administrateur", width=button_width, height=button_height, font=("Arial", 14, "bold"), image=admin_icon_ctk, compound="left", fg_color="#FFFFFF", hover_color="#E6E6FF", text_color="#4C4C4C", border_color="#808080", border_width=1, command=afficher_non_active)

            admin_button_reference = admin_button
            update_element_positions.admin_button_window_id = canvas.create_window(admin_btn_x, admin_btn_y, window=admin_button, tags="admin_button")
        else:
            canvas.coords(update_element_positions.admin_button_window_id, admin_btn_x, admin_btn_y)
        current_y += button_height + button_pady

        acces_btn_x = center_x + (button_height / 0.111)
        acces_btn_y = current_y + (button_height / 0.45)
        if not hasattr(update_element_positions, 'acces_button_window_id'):
            global acces_icon_ctk
            try:
                chemin_acces_icon = os.path.join(dossier_script, "Acces.png")
                acces_icon_ctk = ctk.CTkImage(light_image=Image.open(chemin_acces_icon), dark_image=Image.open(chemin_acces_icon), size=(25, 25))
            except Exception as e:
                print(f"Erreur de chargement de l'icône Acces.png: {e}")
                acces_icon_ctk = None

            acces_button = ctk.CTkButton(canvas, text="Contrôle d'accès", width=button_width, height=button_height, font=("Arial", 14, "bold"), image=acces_icon_ctk, compound="left", fg_color="#FFFFFF", hover_color="#E6E6FF", text_color="#4C4C4C", border_color="#808080", border_width=1, command=lambda: [racine.withdraw(), ouvrir_fenetre_scan()])
            update_element_positions.acces_button_window_id = canvas.create_window(acces_btn_x, acces_btn_y, window=acces_button, tags="acces_button")
        else:
            canvas.coords(update_element_positions.acces_button_window_id, acces_btn_x, acces_btn_y)

    canvas.bind("<Configure>", update_element_positions)
    racine.update_idletasks()
    update_element_positions()

    label_footer = tk.Label(racine, text="Développé par Aoudjit kosseila — Version 1.0", font=("Arial", 10), fg="#555555", bg="#FFFFFF", anchor="w")
    label_footer.place(relx=0.01, rely=0.98, anchor="sw")
    
        # Afficher la fenêtre d'activation au démarrage si non activée
    if not verifier_licence_locale():
        racine.after(500, ouvrir_fenetre_version)

    racine.mainloop()

if __name__ == "__main__":
    ouvrir_choix_utilisateur()